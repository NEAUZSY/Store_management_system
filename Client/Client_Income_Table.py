# 此文件用于生成利润表
import os

import openpyxl as xl
from openpyxl.worksheet.page import PageMargins
from openpyxl.styles import Border, Side, PatternFill, Font, Alignment

from SQL import MyDb
from Client_Get_Income_Table import commodity, Income_Table
from utils.log import log
from utils.generate_pdf import PDFConverter
from utils.upper_value import cnumber

DATA = [['01电子计算机',
         '01*计算机',
         '1',
         '1',
         '个',
         '是',
         '0',
         '0',
         '0',
         '10',
         '11',
         '110',
         '5',
         '20',
         '100',
         '5',
         '11',
         '55',
         '39.82',
         '5.18',
         '0.62',
         '0.063',
         '33.96',
         '39.82',
         '能源']]


def generate_income_table(item_list, date, save_Excle=True):
    data_len = len(item_list)

    wb = xl.Workbook()
    ws = wb.active

    ws.set_printer_settings(paper_size=9, orientation='landscape')
    ws.page_margins = PageMargins(left=0, right=0, top=1, bottom=1)  # 设置页边距(单位:英寸)
    # -------------开始设置格式-------------#
    # 设置缩放比例
    ws.views.sheetView[0].zoomScale = 115
    # 设置打印格式
    ws.print_options.horizontalCentered = True
    # ws.print_options.verticalCentered = True

    # 设置行高
    for i in range(1, data_len + 8):
        ws.row_dimensions[i].height = 20

    # 设置列宽
    col_width = [13, 12, 8, 8, 4.2, 4.2, 4.2, 4.2, 4.2, 4.2, 4.2, 4.2, 4.2, 4.2, 4.2,
                 4.2, 4.2, 4.2, 5.6, 5.6, 4.7, 5.6, 5.6, 4.7, 4.7]
    col = [chr(i) for i in range(ord("A"), ord("Y") + 1)]
    for i, width in enumerate(col_width):
        ws.column_dimensions[col[i]].width = width

    # 合并标题单元格
    ws.merge_cells('A1:Y1')
    ws.merge_cells('A2:Y2')

    # 设置边框 颜色 标题
    color_fill = PatternFill("solid", fgColor="EDEDED")
    titles = ['大类', '二级分类', '名  称', '规格型号', '单位', '是否含税',
              '期初', '', '',
              '购入', '', '',
              '销售', '', '',
              '库存', '', '',
              '毛利', '税费', '', '',
              '净利润', '利润率 %', '往来单位']
    rangeCell = ws['A4:Y{}'.format(6 + data_len)]
    for i, r in enumerate(rangeCell):
        for j, c in enumerate(r):
            c.border = Border(bottom=Side(style='thin', color='000000'),
                              right=Side(style='thin', color='000000'),
                              left=Side(style='thin', color='000000'),
                              top=Side(style='thin', color='000000'))
            if i in (0, 1, data_len + 2):
                c.fill = color_fill
                c.font = Font(size=9)
                c.alignment = Alignment(horizontal='center', vertical='center', wrapText=True)
            else:
                c.font = Font(size=7)
                c.alignment = Alignment(horizontal='center', vertical='center')
            if i == 0:
                c.value = titles[j]
                print(1, j)

    title_2_cells = ws['G5:V5'][0]
    title_2 = ['数量', '单价', '金额',
               '数量', '单价', '金额',
               '数量', '单价', '金额',
               '数量', '单价', '金额',
               '', '增值税', '附加', '印花税']
    for i, cell in enumerate(title_2_cells):
        cell.value = title_2[i]

    # 合并首行表头
    for char in [chr(i) for i in range(ord("A"), ord("F") + 1)]:
        ws.merge_cells('{}4:{}5'.format(char, char))
    ws.merge_cells('S4:S5')
    ws.merge_cells('W4:W5')
    ws.merge_cells('X4:X5')
    ws.merge_cells('Y4:Y5')

    # 合并首行四大项 必须在赋值之后合并
    ws.merge_cells('G4:I4')
    ws.merge_cells('J4:L4')
    ws.merge_cells('M4:O4')
    ws.merge_cells('P4:R4')
    ws.merge_cells('T4:V4')

    ws.merge_cells('A{}:F{}'.format(data_len + 6, data_len + 6))

    title_cell_1 = ws['A1']
    title_cell_1.value = '唐山迈拓科技有限公司'
    title_cell_1.font = Font(size=12)
    title_cell_1.alignment = Alignment(horizontal='center', vertical='center')

    title_cell_2 = ws['A2']
    title_cell_2.value = '{}年{}月利润表'.format(date[:4], "%d" % int(date[5:7]))
    title_cell_2.font = Font(size=16)
    title_cell_2.alignment = Alignment(horizontal='center', vertical='center')

    title_sum = ws['A{}'.format(data_len + 6)]
    title_sum.value = '本月累计'

    log('初始化工作表完成，开始写入数据')

    # -------------开始写入数据-------------#
    value = 0
    for i, item in enumerate(item_list):
        row = ws['A{}:X{}'.format(i + 6, i + 6)][0]
        # info = [i + 1, item[6], item[7], item[9], item[10], item[12], '']
        # value += float(item[12])
        for j, cell in enumerate(row):
            cell.value = item[j]

    log('数据写入完成，开始保存')
    file_name = date + '利润表.xlsx'
    wb.save(file_name)

    log('开始转换PDF')
    path = './' + file_name
    pdfConverter = PDFConverter(path)
    pdfConverter.run_conver()
    if not save_Excle:
        os.remove(path)
    log('出库单生成成功')


def wash(table):
    tt = list()
    for row in table:
        if type(row) is commodity:
            tt.append(row.info[1:])
    return tt


if __name__ == '__main__':
    it = Income_Table('2022-01', MyDb())
    data_list = it.generate_data(it)
    generate_income_table(wash(data_list), '2022-01')
