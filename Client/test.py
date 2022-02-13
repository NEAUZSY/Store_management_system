import os

import openpyxl as xl
from openpyxl.styles import Border, Side, PatternFill, Font, Alignment

from utils.log import log
from utils.generate_pdf import PDFConverter
from utils.upper_value import cnumber

DATA = [
    ('0', '220102344', '2022-02-13 00:00:00', '迈拓', '03计算机配套产品', '20*电源', '电源', '400W', '块', '16.0', '40.71', '36.03',
     '651.33', '576.4', '01963924-35678'), (
        '0', '222100466', '2022-02-13 00:00:00', '迈拓', '02计算机配套产品', '22*键盘', '键盘', '罗技 K120', '个', '1.0', '72.57',
        '64.22',
        '72.57', '64.22', '01963924-35678')]


def generate_odo(item_list, id, user, save_Excle=True):
    data_len = len(item_list)

    wb = xl.Workbook()
    ws = wb.active
    # -------------开始设置格式-------------#
    # 设置缩放比例
    ws.views.sheetView[0].zoomScale = 115
    # 设置打印格式
    ws.print_options.horizontalCentered = True
    # ws.print_options.verticalCentered = True

    # 设置行高
    for i in range(1, data_len + 8):
        ws.row_dimensions[i].height = 27

    # 设置列宽
    col_width = [5.22, 14, 12.44, 6.44, 8.11, 12.22, 20]
    col = ['A', 'B', 'C', 'D', 'E', 'F', 'G']
    for i, width in enumerate(col_width):
        ws.column_dimensions[col[i]].width = width

    # 设置边框 颜色 标题
    rangeCell = ws['A5:G{}'.format(6 + data_len)]
    color_fill = PatternFill("solid", fgColor="EDEDED")
    titles = ['序号', '名  称', '规格型号', '数量', '单价', '金额', '备注']
    for i, r in enumerate(rangeCell):
        for j, c in enumerate(r):
            c.border = Border(bottom=Side(style='thin', color='000000'),
                              right=Side(style='thin', color='000000'),
                              left=Side(style='thin', color='000000'),
                              top=Side(style='thin', color='000000'))
            c.alignment = Alignment(horizontal='center', vertical='center')
            if i in (0, data_len + 1):
                c.fill = color_fill
                c.font = Font(size=11)
            if i == 0:
                c.value = titles[j]

    ws.merge_cells('A1:G1')
    ws.merge_cells('A2:G2')
    ws.merge_cells('A3:G3')
    ws.merge_cells('A4:E4')
    ws.merge_cells('F4:G4')
    ws.merge_cells('A{}:E{}'.format(data_len + 6, data_len + 6))
    ws.merge_cells('A{}:F{}'.format(data_len + 7, data_len + 7))

    title_cell_1 = ws['A1']
    title_cell_1.value = '唐山迈拓科技有限公司'
    title_cell_1.font = Font(size=16)
    title_cell_1.alignment = Alignment(horizontal='center', vertical='center')

    title_cell_2 = ws['A2']
    title_cell_2.value = '出库单'
    title_cell_2.font = Font(size=16)
    title_cell_2.alignment = Alignment(horizontal='center', vertical='center')

    title_date_cell = ws['A3']
    date = item_list[0][2][:10]
    title_date_cell.value = "{}年{}月{}日".format(date[:4], date[5:7], date[8:10])
    title_date_cell.font = Font(size=11)
    title_date_cell.alignment = Alignment(horizontal='center', vertical='center')

    source_cell = ws['A4']
    source_cell.value = '往来单位： {}'.format(item_list[0][3])
    source_cell.alignment = Alignment(horizontal='left', vertical='center')

    id_cell = ws['F4']
    id_cell.value = '单据号：CK {}-{}-{}'.format(date[:4], date[5:7], "%03d" % id)
    id_cell.alignment = Alignment(horizontal='right', vertical='center')

    log('初始化工作表完成，开始写入数据')

    # -------------开始写入数据-------------#
    value = 0
    for i, item in enumerate(item_list):
        row = ws['A{}:G{}'.format(i + 6, i + 6)][0]
        info = [i + 1, item[6], item[7], item[9], item[10], item[12], '']
        value += float(item[12])
        for j, cell in enumerate(row):
            cell.value = info[j]

    value = '%.2f' % value
    pt = cnumber()
    upper_value = pt.cwchange(value)

    value_upper_cell = ws['A{}'.format(data_len + 6)]
    value_upper_cell.value = '合计：金额大写: ' + upper_value

    value_cell = ws['F{}'.format(data_len + 6)]
    value_cell.value = str(value)

    other_info_cell = ws['A{}'.format(data_len + 7)]
    other_info_cell.value = ' 记账：        保管：        验收：        提货人：        制单：'
    other_info_cell.font = Font(size=11)
    other_info_cell.alignment = Alignment(horizontal='left', vertical='center')

    name_cell = ws['G{}'.format(data_len + 7)]
    name_cell.value = user
    name_cell.font = Font(size=11)
    name_cell.alignment = Alignment(horizontal='left', vertical='center')

    # if not item[2][:10] == item_list:
    #     print('所选商品日期不一致，请重新选则')
    #     return
    # if not item[3] == source:
    #     print('所选商品往来单位不一致，请重新选则')
    #     return

    log('数据写入完成，开始保存')
    file_name = item_list[0][3] + date + '.xlsx'
    wb.save(file_name)

    log('开始转换PDF')
    path = './' + file_name
    pdfConverter = PDFConverter(path)
    pdfConverter.run_conver()
    if not save_Excle:
        os.remove(path)
    log('出库单生成成功')


if __name__ == '__main__':
    generate_odo(DATA, 1, '刘朝静')
