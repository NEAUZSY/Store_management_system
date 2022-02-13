# -*- encoding=utf-8 -*-
import os
import time
import tkinter
import openpyxl as xl
from openpyxl.styles import Border, Side, PatternFill, Font, Alignment
from tkinter import *
from tkinter import messagebox
from tkinter import ttk
from multiprocessing import Queue, Process
from utils.log import log
from utils.generate_pdf import PDFConverter
from utils.upper_value import cnumber

from SQL import MyDb

info = [
    ('0', '220102344', '2022-02-13 00:00:00', '迈拓', '03计算机配套产品', '20*电源', '电源', '400W', '块', '16.0', '40.71', '36.03',
     '651.33', '576.4', '01963924-35678'), (
        '0', '222100466', '2022-02-13 00:00:00', '迈拓', '02计算机配套产品', '22*键盘', '键盘', '罗技 K120', '个', '1.0', '72.57',
        '64.22',
        '72.57', '64.22', '01963924-35678')]


class ODO(object):
    def __init__(self, info_, db, User):

        self.User = User

        self.db = db
        self.is_check = True
        self.item_list = []
        self.win = tkinter.Tk()  # 窗口
        self.win.title('出库信息查询查询')  # 标题
        screenwidth = self.win.winfo_screenwidth()  # 屏幕宽度
        screenheight = self.win.winfo_screenheight()  # 屏幕高度
        width = 1400
        height = 600
        x = int((screenwidth - width) / 2)
        y = int((screenheight - height) / 2)
        self.win.geometry('{}x{}+{}+{}'.format(width, height, x, y))  # 大小以及位置

        tabel_frame = tkinter.Frame(self.win)
        tabel_frame.pack()

        xscroll = Scrollbar(tabel_frame, orient=HORIZONTAL)
        yscroll = Scrollbar(tabel_frame, orient=VERTICAL)

        columns = columns = ['生成出库单次数', '序号', '入库时间', '往来单位', '一级分类',
                             '二级分类', '商品名称', '规格型号', '单位',
                             '数量', '单价（含税）', '单价（未税）', '金额（含税）',
                             '金额（未税）', '备注/序列号']
        table = ttk.Treeview(
            master=tabel_frame,  # 父容器
            height=25,  # 表格显示的行数,height行
            columns=columns,  # 显示的列
            show='headings',  # 隐藏首列
            selectmode="extended",  # 选择模式为可多选
            xscrollcommand=xscroll.set,  # x轴滚动条
            yscrollcommand=yscroll.set,  # y轴滚动条
        )
        for column in columns:
            table.heading(column=column, text=column, anchor=CENTER,
                          command=lambda name=column:
                          messagebox.showinfo('', '{}描述信息~~~'.format(name)))  # 定义表头
            table.column(column=column, width=110, minwidth=110, anchor=CENTER, )  # 定义列
        xscroll.config(command=table.xview)
        xscroll.pack(side=BOTTOM, fill=X)
        yscroll.config(command=table.yview)
        yscroll.pack(side=RIGHT, fill=Y)
        for index, data in enumerate(info_):
            data_display = tuple([data[-1]] + list(data[:-2]))
            table.insert('', END, values=data_display)  # 添加数据到末尾

        table.bind('<<TreeviewSelect>>', self.selectTree)
        table.pack(fill=BOTH, expand=True)
        self.table = table

        btn_generate = Button(self.win, text="生成出库单", command=self.Click_generate)
        btn_generate.place(x=100, y=height - 50)

        btn_back = Button(self.win, text="返回", command=self.Click_Back)
        btn_back.place(x=200, y=height - 50)

        self.win.mainloop()

    def selectTree(self, event):
        item_list = []
        for item in self.table.selection():
            item_list.append(self.table.item(item, "values"))
        self.item_list = item_list

    def Click_generate(self):

        tasks = len(self.item_list) + 7  # 此处应该是6
        month = self.item_list[0][2][5:7]
        task = 'SELECT `CK_id` FROM `tb_odo` WHERE `MONTH`={};'.format(month)
        self.db.execute(task)
        ck_id = self.db.cursor.fetchall()
        print(ck_id)

        ck_id = int(ck_id[0][0])

        print(ck_id)

        task = 'UPDATE `tb_odo` SET `CK_id`={} WHERE `MONTH` = {};'.format(ck_id, month)
        self.db.execute(task)

        queue = Queue(5)

        generate_process = Process(target=generate, args=(queue, self.item_list, ck_id, self.User))
        display_process = Process(target=display, args=(queue, tasks,))

        generate_process.start()
        display_process.start()

    def Click_Back(self):
        self.is_check = False
        self.win.destroy()


def generate(q: Queue, item_list, id, user, save_Excle=True):
    log('初始化出库单')
    q.put('初始化出库单')

    data_len = len(item_list)

    wb = xl.Workbook()
    ws = wb.active
    # -------------开始设置格式-------------#
    # 设置缩放比例
    ws.views.sheetView[0].zoomScale = 115
    # 设置打印格式
    ws.print_options.horizontalCentered = True
    # ws.print_options.verticalCentered = True

    # 设置行高
    for i in range(1, data_len + 8):
        ws.row_dimensions[i].height = 27

    # 设置列宽
    col_width = [5.5, 20, 20, 7, 11, 13, 6]
    col = ['A', 'B', 'C', 'D', 'E', 'F', 'G']
    for i, width in enumerate(col_width):
        ws.column_dimensions[col[i]].width = width

    # 设置边框 颜色 标题
    rangeCell = ws['A5:G{}'.format(6 + data_len)]
    color_fill = PatternFill("solid", fgColor="EDEDED")
    titles = ['序号', '名  称', '规格型号', '数量', '单价', '金额', '备注']
    for i, r in enumerate(rangeCell):
        for j, c in enumerate(r):
            c.border = Border(bottom=Side(style='thin', color='000000'),
                              right=Side(style='thin', color='000000'),
                              left=Side(style='thin', color='000000'),
                              top=Side(style='thin', color='000000'))
            if j in (1, 2) and i != 0:
                c.alignment = Alignment(horizontal='left', vertical='center', indent=1)
            elif j in (4, 5) and i != 0:
                c.alignment = Alignment(horizontal='right', vertical='center')
            else:
                c.alignment = Alignment(horizontal='center', vertical='center')

            if i in (0, data_len + 1):
                c.fill = color_fill
                c.font = Font(size=10)
            else:
                c.font = Font(size=11)
            if i == 0:
                c.value = titles[j]

    ws.merge_cells('A1:G1')
    ws.merge_cells('A2:G2')
    ws.merge_cells('A3:G3')
    ws.merge_cells('A4:D4')
    ws.merge_cells('E4:G4')
    ws.merge_cells('A{}:E{}'.format(data_len + 6, data_len + 6))
    ws.merge_cells('A{}:G{}'.format(data_len + 7, data_len + 7))

    title_cell_1 = ws['A1']
    title_cell_1.value = '唐山迈拓科技有限公司'
    title_cell_1.font = Font(size=16)
    title_cell_1.alignment = Alignment(horizontal='center', vertical='center')

    title_cell_2 = ws['A2']
    title_cell_2.value = '出库单'
    title_cell_2.font = Font(size=16)
    title_cell_2.alignment = Alignment(horizontal='center', vertical='center')

    title_date_cell = ws['A3']
    date = item_list[0][2][:10]
    title_date_cell.value = "{}年{}月{}日".format(date[:4], date[5:7], date[8:10])
    title_date_cell.font = Font(size=11)
    title_date_cell.alignment = Alignment(horizontal='center', vertical='center')

    source_cell = ws['A4']
    source_cell.value = '往来单位： {}'.format(item_list[0][3])
    source_cell.alignment = Alignment(horizontal='left', vertical='center')

    id_cell = ws['E4']
    id_cell.value = '单据号：CK {}{}-{}'.format(date[:4], date[5:7], "%03d" % id)
    id_cell.alignment = Alignment(horizontal='right', vertical='center')

    log('初始化工作表完成，开始写入数据')
    q.put('初始化工作表完成，开始写入数据')
    # -------------开始写入数据-------------#
    db = MyDb()
    value = 0
    for i, item in enumerate(item_list):
        row = ws['A{}:G{}'.format(i + 6, i + 6)][0]
        info_ = [i + 1, item[6], item[7], item[9], item[10], item[12], '']
        value += float(item[12])
        for j, cell in enumerate(row):
            if info_[j] == 'None':
                info_[j] = ' '
            cell.value = info_[j]

        times = int(item[0]) + 1
        task = 'UPDATE tb_sell SET `generate_times`={} WHERE id={};'.format(times, item[1])
        db.execute(task)
        time.sleep(0.1)
        q.put('正在处理第{}条信息'.format(i))

    q.put('写入其他公共信息')
    value = '%.2f' % value
    pt = cnumber()
    upper_value = pt.cwchange(value)

    value_upper_cell = ws['A{}'.format(data_len + 6)]
    value_upper_cell.value = '合计金额大写: ' + upper_value

    value_cell = ws['F{}'.format(data_len + 6)]
    value_cell.value = '￥' + str(value)

    other_info_cell = ws['A{}'.format(data_len + 7)]
    other_info_cell.value = ' 记账：        保管：        验收：        提货人：        制单： %s' % user
    other_info_cell.font = Font(size=11)
    other_info_cell.alignment = Alignment(horizontal='center', vertical='center')

    log('数据写入完成，开始保存')
    q.put('数据写入完成，开始保存')
    file_name = item_list[0][3] + date + '.xlsx'
    wb.save(file_name)

    log('开始转换PDF')
    q.put('开始转换PDF')
    path = './' + file_name
    pdfConverter = PDFConverter(path)
    pdfConverter.run_conver()
    if not save_Excle:
        os.remove(path)
    log('出库单生成成功')
    q.put('出库单生成成功')


def display(q: Queue, pages):
    print('创建进度窗口')
    root = Tk()
    root.title('执行进度')
    screenwidth = root.winfo_screenwidth()  # 屏幕宽度
    screenheight = root.winfo_screenheight()  # 屏幕高度
    width = 300
    height = 100
    x = int((screenwidth - width) / 2)
    y = int((screenheight - height) / 2)
    root.geometry('{}x{}+{}+{}'.format(width, height, x, y))  # 大小以及位置

    # 当前任务信息框
    process = StringVar(value='正在执行:')
    Label(root, textvariable=process).place(x=20, y=20)

    # 进度条初始化
    Label(root, text='当前进度:', ).place(x=20, y=50)
    progressbar = ttk.Progressbar(root, length=200)
    progressbar['maximum'] = pages
    progressbar.place(x=85, y=50)

    print('开始更新进度信息')
    tasks = []
    while True:
        try:
            task = q.get(timeout=2)
            tasks.append(task)
            progressbar['value'] = len(tasks)
            process.set('正在执行: %s' % task)
            root.update()
        except Exception as e:
            print(e)
            print('全部保存完毕')
            break
    root.destroy()
    print('关闭进度窗口')


def main(info_):
    quare = ODO(info_, MyDb())


if __name__ == '__main__':
    main(info)
